from numpy import empty
from . import prisma

import pandas as pd
from openpyxl import load_workbook
import os
import datetime
import json


def to_excel(dst: str, table: pd.DataFrame, append: bool = False):
    """
    Exporta a una tabla de excel. Puede escribir al final de una ya creada también.

    :param dst: Ruta de destino. Debe contener el nombre del archivo en ella. P.e.: "./excel.xlsx"
    :param table: Tabla a exportar a excel
    :param append: Append to an already existing table.
    """

    if append and os.path.exists(dst):
        wb = load_workbook(dst)
        sheet = wb.worksheets[0]

        table.to_excel(dst, engine='openpyxl', startrow=sheet.max_row, index = False, header= False)   
    
    else:
        table.to_excel(dst, engine='openpyxl')


def cloud_accounts(prismasession) -> pd.DataFrame:
    """
    Obtiene un DataFrame con todas las cuentas del tenant.

    :param prismasession: datos de la sesión con el tenant (url, token)
    :return: DataFrame con cada una de las cuentas del tenant
    """
    
    res = prisma.list_cloud_accounts(prismasession)
    return pd.json_normalize(res)


def cloud_org_accounts(prismasession, cloud_type: str, id: str) -> pd.DataFrame:
    """
    Obtiene un DataFrame con todas las cuentas hijas de una organización.

    :param prismasession: datos de la sesión con el tenant (url, token)
    :param cloud_type: tipo de nube
    :param id: ID de la cuenta organización
    :return: DataFrame con cada una de las cuentas hijas de una organización
    """

    _cloud_type = cloud_type.lower()
    if (_cloud_type == 'aws' or _cloud_type == 'azure' or _cloud_type == 'gcp' or _cloud_type == 'oci'):
        res = prisma.list_cloud_org_accounts(prismasession, _cloud_type, id)
        return pd.json_normalize(res)
    else:
        raise Exception('Error introduciendo tipo de nube. Valores aceptables: aws, azure, gcp, oci.')


def account_groups(prismasession, detailed=True) -> pd.DataFrame:
    """
    Obtiene un DataFrame con todas las account groups del tenant.
    
    :param detailed: True si queremos info detallada.
    :param prismasession: datos de la sesión con el tenant (url, token).
    :return: DataFrame con cada una de las account group del tenant.
    """
    
    res = prisma.list_account_groups(prismasession, detailed)
    return pd.json_normalize(res)

def compliance_sumary(prismasession, framework) -> pd.DataFrame:
    """
    Obtiene un DataFrame con el resumen de compliance de un framework de Prisma.

    Obtiene: failedResources, passedResources, totalResources, highSeverityFailedResources,
    mediumSeverityFailedResources y lowSeverityFailedResources.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param framework: framework del que se busca el compliance
    :return: DataFrame con el resumen compliance.
    """
    
    standards = frameworks(prismasession)
    selected_id = standards.loc[standards['name'] == framework].head(1).iloc[0]['id']


    res = prisma.get_standard_compliance_statistics(prismasession, selected_id)
    return pd.json_normalize(res["summary"])


def compliance_details(prismasession, framework) -> pd.DataFrame:
    """
    Obtiene un DataFrame con el compliance de cada resourceType de un framework de Prisma.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param framework: framework del que se busca el compliance
    :return: DataFrame con compliance detallado por resource de un framework de Prisma
    """
    
    standards = frameworks(prismasession)
    selected_id = standards.loc[standards['name'] == framework].head(1).iloc[0]['id']


    rawResponse = prisma.get_standard_compliance_statistics(prismasession, selected_id)
    
    res = []
    for summary in rawResponse["requirementSummaries"]:
        for section in summary["sectionSummaries"]:
            res.append(section)

    return pd.json_normalize(res)


def historic(prismasession, framework) -> pd.DataFrame:
    """
    Obtiene un DataFrame con el historial de compliance de un framework de Prisma.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param framework: framework del que se busca el compliance
    :return: DataFrame con el historial de compliance de un framework de Prisma
    """
    
    standards = frameworks(prismasession)
    selected_id = standards.loc[standards['name'] == framework].head(1).iloc[0]['id']

    res = prisma.get_standard_compliance_statistics_trend(prismasession, selected_id)
    res = pd.json_normalize(res)
    res['date'] = res.apply(lambda x: timestamp_to_date(x), 1)

    return res


def timestamp_to_date(x):
    _x = json.loads(x.to_json())
    date = datetime.datetime.fromtimestamp(_x['timestamp']/1000).strftime("%d/%m/%Y")
    return date


def frameworks(prismasession) -> pd.DataFrame:
    """
    Obtiene un DataFrame con todos los frameworks (compliance standards) de Prisma.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :return: DataFrame con cada uno de los standards de Prisma
    """
    
    res = prisma.list_compliance_standards(prismasession)
    return pd.json_normalize(res)


def framework_requirements(prismasession, framework_name) -> pd.DataFrame:
    """
    Obtiene todos los requerimientos de un framework de Prisma.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param framework_name: nombre del framework
    :return: DataFrame con cada uno de los standards de Prisma
    """
    
    standards = frameworks(prismasession)
    try:
        selected_id = standards.loc[standards['name'] == framework_name].head(1).iloc[0]['id']
    except:
        raise Exception("Error obteniendo requiermientos. Probablemente, nombre de framework incorrecto.")

    res = prisma.list_compliance_requirements(prismasession, selected_id)
    return pd.json_normalize(res)


def framework_sections(prismasession, framework_name, requirement_name) -> pd.DataFrame:
    """
    Obtiene todas las secciones de un framework de Prisma.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param framework_name: nombre del framework
    :param requirement_name: nombre del requerimiento
    :return: DataFrame con cada uno de las secciones de un requerimiento de Prisma
    """

    reqs = framework_requirements(prismasession, framework_name)
    try:
        req_id = reqs.loc[reqs['name'] == requirement_name].head(1).iloc[0]['id']
    except:
        raise Exception("Error obteniendo requiermientos. Probablemente, nombre de framework incorrecto.")

    res = prisma.list_compliance_sections(prismasession, req_id)

    return pd.json_normalize(res)


def resources_usage_over_time(prismaSession, accounts_ids, start_date: str = None,
                              end_date: str = None) -> pd.DataFrame:
    """
    Obtiene un json con el uso de recursos licenciables sobre el tiempo. Si no se concreta fecha de inicio o fecha de final,
    se asume que se busca sin acotación de fechas. Para acotar, hay que definir AMBAS fechas, inicio Y final.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param accounts_ids: Lista con IDs de las suscripciones sobre las que obtener el uso
    :param start_date: Fecha de inicio, en formato "dd/mm/yyyy". Por defecto, es el origen
    :param end_date: Fecha de fin, en formato "dd/mm/yyyy". Por defecto, es hoy
    :return: DataFrame con el uso de recursos licenciables sobre el tiempo
    """

    time_range = None
    if (start_date != None and end_date != None):
        time_range = {
            "type": "absolute",
            "value": {
                "startTime": int(datetime.datetime.strptime(start_date, '%d/%m/%Y').timestamp() * 1000),
                "endTime": int(datetime.datetime.strptime(start_date, '%d/%m/%Y').timestamp() * 1000)
            }
        }

    res = prisma.resource_usage_over_time(prismaSession, accounts_ids, time_range)
    return pd.json_normalize(res)


def list_policies(prismaSession, framework) -> pd.DataFrame:
    """
    Obtiene un DataFrame con un listado de politicas para un framework.

    La API de prisma no permite esta funcionalidad por lo que hacemos la inversa,
    se listan las politicas y comprobamos si el framework aparece en sus detalles,
    pues una policy puede pertenecer a varios frameworlks.

    :param prismaSession: datos de la sesión con el tenant (url, token)
    :param framework: nombre del framework de las politicas a buscar
    :return: DataFrame con un listado de politicas.
    """

    res = prisma.list_policies(prismaSession)
    res = pd.json_normalize(res)
    res['framework'] = res.apply(lambda x: list_policies_get_standard(x, framework), 1)
    res = res.loc[res['framework'] == framework]
    return res

# Con esta función podemos comprobar si la policy pertenece a un determinado framework
# TODO requirement
def list_policies_get_standard(x, framework):
    _x = json.loads(x.to_json())

    if ('complianceMetadata' in _x.keys() and _x['complianceMetadata'] != None):
        for c in _x['complianceMetadata']:
            if (c['standardName'] == framework):
                return framework
        return ''
    else:
        return ''


def alerts(prismasession: prisma.PrismaSession, excelPath: str, timeType: str = 'relative', timeAmount: str = '30', 
            timeUnit: prisma.TimeUnit = prisma.TimeUnit.DAY,
           detailed: bool = True, sortBy: str = None, offset: int = 0, pageToken: str = None,
           alertId: str = None, alertStatus: prisma.AlertStatus = None, cloudAccount: str = None, cloudAccountId: str = None,
           accountGroup: str = None,
           cloudType: str = None, cloudRegion: str = None, cloudService: str = None, policyId: str = None,
           policyName: str = None,
           policySeverity: prisma.PolicySeverity = None, policyLabel: str = None, policyType: prisma.PolicyType = None,
           policyFramework: str = None,
           policyComplianceRequirement: str = None, policyComplianceSection: str = None, policyRemediable: bool = None,
           alertRuleName: str = None,
           resourceId: str = None, resourceName: str = None, resourceType: str = None):
    """
    Obtiene un DataFrame de alertas de Prisma paginada.

    :param prismaSession: datos de la sesión con el tenant (url, token).
    :param timeType: Tipo de tiempo, default 'relative'.
    :param timeAmount: Cantidad de tiempo, unidad definida en el parámetro timeUnit, default 30.
    :param timeUnit: Unidad de tiempo, default 'year'.
    :param detailed: Alerta detallada o no, default 'True'.
    :param sortBy: Ordenar las alertas. El formato del parámetro es PROPERTY:asc para ascendiente, PROPERTY:desc para descendiente. P.e.: sortBy=id:desc, sortBy=firstseen:asc, sortBy=lastseen:desc).
        Propiedades válidas: firstseen, lastseen, resource.regionid, alerttime, id, resource.accountid, status y resource.id.
    :param offset: Número de alertas que saltar (ignorar) en los resultados.
    :param limit: Número máximo de alertas, no más de 10000. Default 10000.
    :param pageToken: Identificador de la página de alertas, para cuando las alertas no caben en una respuesta.
    :param alertId: Alert ID.
    :param alertStatus: Enum AlertStatus.
    :param cloudAccount: Cloud account.
    :param cloudAccountId: ID de la cloud account.
    :param accountGroup: Account group.
    :param cloudType: Cloud type, enum CloudType.
    :param cloudRegion: Cloud region.
    :param cloudService: Cloud service.
    :param policyId: ID de la policy.
    :param policyName: Nombre de la policy.
    :param policySeverity: PolicySeverity enum.
    :param policyLabel: Label de la policy.
    :param policyType: PolicyType enum.
    :param policyFramework: Nombre del Framework.
    :param policyComplianceRequirement: Nombre del Compliance Requirement.
    :param policyComplianceSection: ID del Compliance Section.
    :param policyRemediable: Bool. True es remediable, False no.
    :param alertRuleName: Nombre de la Alert Rule.
    :param resourceId: ID del resource.
    :param resourceName: Nombre del resource.
    :param resourceType: Tipo del resource.
    :return: DataFrame con las alertas del tenant
    """

    # Prisma devuelve muchísimos metadatos, vamos a quedarnos con los interesantes
    def transform_response(response: json) -> json:
        
        res = []
        for item in response['items']:
            res.append({
                    'alertId': item['id'],
                    'alertStatus': item['status'],
                    'alertTime': item['alertTime'],
                    'resourceType': item['resource']['resourceType'],
                    'cloudType': item['resource']['cloudType'],
                    'accountName': item['resource']['account'],
                    'policyId': item['policyId']
            })
        return res

    
    res = prisma.list_alerts_v2(prismasession, timeType=timeType, timeAmount=timeAmount,
                                         timeUnit=timeUnit, detailed=detailed, fields=None,
                                         sortBy=sortBy, offset=offset, limit=50,
                                         pageToken=pageToken,
                                         alertId=alertId, alertStatus=alertStatus, cloudAccount=cloudAccount,
                                         cloudAccountId=cloudAccountId, accountGroup=accountGroup,
                                         cloudType=cloudType, cloudRegion=cloudRegion, cloudService=cloudService,
                                         policyId=policyId, policyName=policyName,
                                         policySeverity=policySeverity, policyLabel=policyLabel,
                                         policyType=policyType,
                                         policyComplianceStandard=policyFramework,
                                         policyComplianceRequirement=policyComplianceRequirement,
                                         policyComplianceSection=policyComplianceSection,
                                         policyRemediable=policyRemediable,
                                         alertRuleName=alertRuleName, resourceId=resourceId,
                                         resourceName=resourceName,
                                         resourceType=resourceType)
    try:
        nextPageToken = res['nextPageToken']
    
    except KeyError:    
        res = pd.json_normalize(transform_response(res))
        if not res.empty:
            to_excel(excelPath, res, True)
    
    else:
        res = pd.json_normalize(transform_response(res))
        res.count()
        to_excel(excelPath, res, os.path.exists(excelPath))
        alerts(prismasession, excelPath, pageToken=nextPageToken)

    
    


def login(apiUrl: str = None, access_key_id: str = None, access_key_pass: str = None):
    """
    Login en la api de prisma para empezar a hacer querys.

    :param apiUrl: Cluster donde se encuentra Prisma
    :param access_key_id: Access Key ID
    :param access_key_pass: Access Key Pass
    :return: Objeto con los datos de sesión (url, token)
    """
    
    ps = prisma.prisma_login(apiUrl, access_key_id, access_key_pass)
    return ps
