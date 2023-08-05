"""
a small tool to assign students(workers) to projects(jobs)
based on their rank of preference(cost) to each of the projects(jobs)

Two functions, *assign_project()* and *generate_sample()*, realize
the facility.
"""

from munkres import Munkres
import csv
import openpyxl
from datetime import datetime
import random


def assign_project(csv_name, capacity, **kwargs):
    """This function is for assigning.

    :param csv_name: The name of the input csv file.
    :param capacity: the number of students that a project can take at most.
    :param kwargs: **argument. Can take "output", "time" and "deviation" 3 fields. Please see README file for details
    :return: None
    """

    start_time = datetime.now()
    matrix = __convert_csv(csv_name, capacity)
    student_num = len(matrix)
    project_num = len(matrix[0]) // capacity

    if "output" in kwargs and kwargs["output"] == "file":
        result = openpyxl.Workbook()
        result.create_sheet("Sheet1")
        result_sheet_s = result["Sheet1"]
        result_sheet_s.append(["Student", "Project", "Choice"])
        result.create_sheet("Sheet2")
        result_sheet_p = result["Sheet2"]
        result_sheet_p.append(["PID", "Student1", "Student2", "Student3", "Student4", "Student5"])

    m = Munkres()
    indexes = m.compute(matrix)

    assigned_choices = []
    project_assigned_people = [0] * project_num

    for row, column in indexes:
        value = matrix[row][column]

        project_assigned_people[column % project_num] += 1

        if "output" not in kwargs or kwargs["output"] == "stdout":
            print(f'({row}, {column}) -> {value}')
            print(
                f"Student {row + 1} is assigned to project {column % project_num + 1}, this is his/her #{value} choice")

        if "output" in kwargs and kwargs["output"] == "file":
            result_sheet_s.append([row + 1, column % project_num + 1, value])
            if project_assigned_people[column % project_num] == 1:
                result_sheet_p.cell(column % project_num + 2, 1).value = column % project_num + 1
            result_sheet_p.cell(column % project_num + 2,
                                project_assigned_people[column % project_num] + 1).value = row + 1

        assigned_choices.append(value)

    if "deviation" not in kwargs or kwargs["deviation"]:
        print(f"The standard deviation of the assigned choices is {__calculate_s(assigned_choices)}")
    end_time = datetime.now()

    if "time" not in kwargs or kwargs["time"]:
        print(f"The total time used is {end_time - start_time}")

    if "output" in kwargs and kwargs["output"] == "file":
        for i in range(project_num):
            if project_assigned_people[i] == 0:
                result_sheet_p.cell(i + 2, 1).value = i + 1

        result.save("result.xlsx")


def __convert_csv(csv_name, capacity):
    with open(csv_name, encoding='UTF-8-sig') as source:
        reader = csv.reader(source)
        matrix = []
        for row in reader:
            line = []
            for number in row:
                if number == '' or number == ' ':
                    line.append(20)
                else:
                    line.append(int(number))
            line *= capacity
            matrix.append(line)
    return matrix


def __calculate_s(int_list):
    total = 0
    for number in int_list:
        total += number
    avg = total / len(int_list)
    result = 0
    for number in int_list:
        result += (number - avg) ** 2
    result /= len(int_list) - 1
    result **= 0.5
    return result


def generate_sample(csv_name, student_num, project_num):
    """

    :param csv_name: The name of the output csv file.
    :param student_num:
    :param project_num:
    :return: None
    """
    with open(csv_name, "w") as target:
        writer = csv.writer(target)
        for i in range(1, student_num + 1):
            index = list(range(1, project_num + 1))
            random.shuffle(index)
            writer.writerow(index)
    print("New Sample generated...")
