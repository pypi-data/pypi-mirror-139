from tkinter import Tk, ttk, filedialog, Entry, Label, StringVar
import openpyxl
from bzmicrostitcher.ProcessData import Stitch, Create_OMETIFF

gui = Tk()
gui.geometry("325x80")
gui.title("Keyence Stitching / OME-TIFF Generator")
gui.resizable(False,False)

def getMainDirectory():
    folder_selected = filedialog.askdirectory()
    MainDirectory_Var.set(folder_selected)

def getXLSX():
    folder_selected = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    OME_XLSX_Var.set(folder_selected)

def StartProcessing():
    OME_XLSX = OME_XLSX_Var.get()
    MainDirectory = MainDirectory_Var.get()
    WSI_List = openpyxl.load_workbook(OME_XLSX).worksheets[0]
    SlideID = []
    GroupID = []
    XYID = []
    GCI_List = []
    ImgPath_List = []
    Path_List = []
    Num_Slides = WSI_List.max_row
    ErrorLog = open(MainDirectory + '/' + 'Error_Log.txt', 'a+')

    for X in range (1, Num_Slides):
        SlideID_Val = WSI_List.cell(row = X+1, column = 1).value
        SlideID.append(SlideID_Val)
        GroupID_Val = WSI_List.cell(row = X+1, column = 2).value
        GroupID.append(GroupID_Val)
        XYID_Val = WSI_List.cell(row = X+1, column = 3).value
        XYID.append(XYID_Val)
        
    for Y in range (0, Num_Slides-1):
        GCI_List.append(MainDirectory + '/' + str(GroupID[Y]) + '/' +
                        str(XYID[Y]) + '/' + '*.gci')
        
        ImgPath_List.append(MainDirectory + '/' + str(GroupID[Y]) + '/' +
                            str(XYID[Y]) + '/' + str(XYID[Y]) + '.tif') 
        
        Path_List.append(MainDirectory + '/' + str(GroupID[Y]) + '/' + 
                            str(XYID[Y]))

    print("Stitching Files")
    for X in range (0, Num_Slides-1):
        try:
            Stitch_arr = Stitch(MainDirectory, GCI_List[X], Path_List[X], ImgPath_List[X], SlideID[X])
            Stitch_arr
        except:
            ErrorLog.write("Something is not right, unable to stitch: " + GCI_List[X] + ", " + 
                  ImgPath_List[X] + ", " + SlideID [X] + "\n")
            pass
    ErrorLog.write("If this is the only line you see then everything went well! YAY!")
    ErrorLog.close()


#Main Directory Selection
MainDirectory_Var = StringVar()
MainLabel = Label(gui ,text="Choose Directory")
MainLabel.grid(row=0,column = 0)
Direc_Entry = Entry(gui,textvariable=MainDirectory_Var)
Direc_Entry.grid(row=0,column=1) 
btnFind = ttk.Button(gui, text="Browse Folder",command = getMainDirectory)
btnFind.grid(row=0,column=2)

#XLSX File Selection
OME_XLSX_Var = StringVar()
OMELabel = Label(gui ,text="Choose XLSX File")
OMELabel.grid(row=1,column = 0)
XLSX_Entry = Entry(gui,textvariable=OME_XLSX_Var)
XLSX_Entry.grid(row=1,column=1)
btnXLSX = ttk.Button(gui, text="Browse Files",command = getXLSX)
btnXLSX.grid(row=1,column=2)

c = ttk.Button(gui, text="Start Stitching", command=StartProcessing)
c.grid(row=4,column=1)
gui.mainloop()