import json
from datetime import datetime
import datetime
from os.path import exists
import os

from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import argparse
import mysql.connector
import yaml

__version__ = "1.3.5"

def main(args=None):
	ruta = os.path.dirname(os.path.abspath(__file__))
	rutaJson = ruta+"/dadesPandora.json"
	parser = argparse.ArgumentParser(description='Una API per a recullir informacio de la web de PandoraFMS.')
	parser.add_argument('-e', '--excel', help='Guardar la informacio a un excel, per defecte esta desactivat', action="store_true")
	parser.add_argument('-f', '--file', help='La ruta(fitxer inclos) a on guardar el excel. Per defecte es: PandoraResum.xlsx', default="PandoraResum.xlsx", metavar="RUTA")
	parser.add_argument('--json-file', help='La ruta(fitxer inclos) a on es guardara el fitxer de dades json. Per defecte es:'+rutaJson, default=rutaJson, metavar='RUTA')
	parser.add_argument('-q', '--quiet', help='Nomes mostra els errors i el missatge de acabada per pantalla.', action="store_false")
	parser.add_argument('-v', '--versio', help='Mostra la versio', action='version', version='PandoraFMS_API-NPP v'+__version__)
	

	conf = ruta+"/config/config.yaml"
	if not(os.path.exists(ruta+"/config")):
		os.mkdir(ruta+"/config")
	if not(os.path.exists(ruta+"/errorLogs")):
		os.mkdir(ruta+"/errorLogs")
	if not(exists(conf)):
		print("Emplena el fitxer de configuracio de Base de Dades a config/config.yaml")
		article_info = [
			{
				'BD': {
				'host' : 'localhost',
				'user': 'root',
				'passwd': 'patata'
				}
			}
		]
		with open(conf, 'w') as yamlfile:
			data = yaml.dump(article_info, yamlfile)

	with open(conf, "r") as yamlfile:
		data = yaml.load(yamlfile, Loader=yaml.FullLoader)

	servidor = data[0]['BD']['host']
	usuari = data[0]['BD']['user']
	contrassenya = data[0]['BD']['passwd']

	try:
		mydb =mysql.connector.connect(
			host=servidor,
			user=usuari,
			password=contrassenya,
			database="pandora"
			)
		mycursor = mydb.cursor(buffered=True)
		print("Access MySQL correcte")
	except:
		try:
			mydb =mysql.connector.connect(
				host=servidor,
				user=usuari,
				password=contrassenya
				)
			print("Base de dades no existeix, creant-la ...")
			mycursor = mydb.cursor(buffered=True)
			mycursor.execute("CREATE DATABASE pandora")
			mydb =mysql.connector.connect(
				host=servidor,
				user=usuari,
				password=contrassenya,
				database="pandora"
				)
			mycursor = mydb.cursor(buffered=True)
			mycursor.execute("CREATE TABLE credencials (usuari VARCHAR(255), contassenya VARCHAR(255), apipassw VARCHAR(255), host VARCHAR(255));")
		except:
			print("Login MySQL incorrecte o MySQL no instal·lat")
			return

	mycursor.execute("SELECT * FROM credencials")
	resultatbd = mycursor.fetchall()
	parser.add_argument('-w', '--web', help='Especificar la web de PandoraFMS a on accedir.', default=resultatbd[0][3], metavar="URL")
	args = parser.parse_args(args)

	url = args.web
	apipassw = resultatbd[0][2]
	user = resultatbd[0][0]
	passwd = resultatbd[0][1]


	metode = "all_agents"
	other = ";|%20|type_row,group_id,agent_name"
	other2 = "url_encode_separator_|"
	try:
		parameters = {"op":"get", "op2":metode, "return_type":"json", "apipass":apipassw, "user":user, "pass":passwd}
		agentsFull = requests.get(url, params=parameters).json()
		metode = "tree_agents"
		parameters = {"op":"get", "op2":metode, "return_type":"json", "other":other, "other_mode":other2, "apipass":apipassw, "user":user, "pass":passwd}
		treeAgents = requests.get(url, params=parameters).json() 
	except Exception as e:
		print("Error de conexio")
		now = datetime.datetime.now()
		date_string = now.strftime('%Y-%m-%d--%H-%M-%S-Connexio')
		f = open(ruta+"/errorLogs/"+date_string+".txt",'w')
		f.write("Error de conexio"+str(e))
		f.close()

	class grup:
		def __init__(self, codi, nom):
			self.codi = codi
			self.nom = nom
	class agent:
		def __init__(self, codi, nom, codiP, status):
			self.codi = codi
			self.nom = nom
			self.codiP = codiP
			self.status = status


	def statusConvertor(status):
		if status == 0:
			return("Correcte")
		elif status == 2:
			return("Warning")
		elif status == 200:
			return("ERROR")
		elif status == 100:
			return("ERROR")
		elif status == 3:
			return("Desconegut/desconectat")
		else:
			return("codi desconegut")

	fitxer = args.file
	if exists(fitxer) == False and args.excel:
		workbook = Workbook()
		workbook.save(fitxer)
	if args.excel:
		workbook = load_workbook(filename = fitxer)

	llistaAgents = []
	llistaGrups = [{"grupID": 0, "SenseGrup" : []}]

	def escriptor(workbook, z, y, grup, agent): #z contadorAgent | y contadorGrup
		if args.excel:
			wsdefault = workbook['Sheet']
		if z == 0 and args.excel:
			wsdefault.cell(row=z+1, column=2*y+1, value=grup.nom)
			wsdefault.cell(row=z+1, column=2*y+2, value="Status")
		llistaAgents.append({"nom":agent.nom, "status":agent.status, "id":agent.codi, "CodiGrup":agent.codiP})
		if args.excel:
			wsdefault.cell(row=z+2, column=2*y+1, value=agent.nom)
			wsdefault.cell(row=z+2, column=2*y+2, value=agent.status)


	try:

		i=0
		num_grups=0
		quantGrups = len(treeAgents['data']) #num total grups
		while i < quantGrups: #crea una variable per cada grupamb la seva info a dins
			if (treeAgents['data'][i]['type_row']) == "group":
				vars()["grup"+str(num_grups)] = grup(treeAgents['data'][i]['group_id'], treeAgents['data'][i]['group_name'])
				if args.quiet:
					print()
				num_grups += 1
			i += 1

		
		i=0
		num_agents=0
		quantAgents = len(treeAgents['data']) #num total agents
		while i < quantAgents: #crea una variable per cada agent amb la seva info a dins
			if (treeAgents['data'][i]['type_row']) == "agent":
				x=0
				while x < len(agentsFull['data']):
					if (treeAgents['data'][i]['agent_id'])==(agentsFull['data'][x]['id_agente']):
						break
					x += 1
				vars()["agent"+str(num_agents)] = agent(treeAgents['data'][i]['agent_id'], treeAgents['data'][i]['alias'], treeAgents['data'][i]['agent_id_group'], statusConvertor(agentsFull['data'][x]['status']))
				num_agents += 1
			i += 1
		y = 0
		contadorGrup = 0

		while y < num_grups:
			if args.quiet:
				print(vars()["grup"+str(y)].nom)
				print("=========================")
			z = 0
			contadorAgent = 0
			while z < num_agents:
				if (vars()["agent"+str(z)].codiP) == (vars()["grup"+str(y)].codi):
					if args.quiet:
						print(vars()["agent"+str(z)].nom)
						print(vars()["agent"+str(z)].status)
					if args.excel:
						escriptor(workbook, contadorAgent, contadorGrup, vars()["grup"+str(y)], vars()["agent"+str(z)])
					else:
						escriptor("PandoraResum.xlsx", contadorAgent, contadorGrup, vars()["grup"+str(y)], vars()["agent"+str(z)])
					if args.excel:
						workbook.save(fitxer)
					if args.quiet:
						print()
					conteAgents = True
					contadorAgent += 1
				z+=1
			if args.quiet:
				print()
			if conteAgents == True:
				contadorGrup += 1
				llistaGrups.append({"grupID": vars()["grup"+str(y)].codi,vars()["grup"+str(y)].nom : llistaAgents})
				llistaAgents = []
			y+=1
	except Exception as e:
		if i==0:
			print("Error de resposta")
			now = datetime.datetime.now()
			date_string = now.strftime('%Y-%m-%d--%H-%M-%S-resposta')
			f = open(ruta+"/errorLogs/"+date_string+".txt",'w')
			f.write("Error de resposta "+ str(e) +treeAgents)
			f.close()
		else:
			print("Error de resposta")
			now = datetime.datetime.now()
			date_string = now.strftime('%Y-%m-%d--%H-%M-%S-resposta')
			f = open(ruta+"/errorLogs/"+date_string+".txt",'w')
			f.write("Error de resposta "+str(e)+ str(agentsFull))
			f.close()
		print(e)

	myList = [{"grups" : llistaGrups}]
	try:
		with open(args.json_file, 'w') as f:
			json.dump(myList, f, indent = 4)
	except Exception as e:
			print("Error d'escriptura de json")
			now = datetime.datetime.now()
			date_string = now.strftime('%Y-%m-%d--%H-%M-%S-json')
			f = open(ruta+"/errorLogs/"+date_string+".txt",'w')
			f.write("Error d'escriptura de json "+str(e))
			f.close()
	if not(args.quiet):
		print("Done")
	
if __name__ =='__main__':
    main()